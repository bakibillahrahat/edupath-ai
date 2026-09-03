from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

_HEADER_FONT = Font(bold=True)


def _write_sheet(sheet: Worksheet, headers: list[str], rows: list[list]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = _HEADER_FONT
    for row in rows:
        sheet.append(row)
    for column_cells in sheet.columns:
        length = max((len(str(cell.value)) for cell in column_cells if cell.value is not None), default=10)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(length + 2, 12), 60)


def build_workflow_workbook(
    *,
    candidate_opportunities: list[dict],
    eligibility_verdicts: list[dict],
    research_match_verdicts: list[dict],
    ranked_opportunities: list[dict],
) -> bytes:
    """Builds the 5-sheet export workbook (spec: Universities, Professors,
    Funding, Eligibility, Ranked Opportunities). Every row is sourced
    directly from real workflow output -- nothing here is fabricated."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    candidates_by_id = {c["id"]: c for c in candidate_opportunities}

    # Sheets are independent views, not mutually-exclusive buckets -- a
    # candidate with both a university and funding attached legitimately
    # appears in both the Universities and Funding sheets.
    universities_sheet = workbook.create_sheet("Universities")
    _write_sheet(
        universities_sheet,
        ["Title", "University", "Program", "Degree Level", "Country", "Official URL"],
        [
            [c.get("title"), c.get("university"), c.get("program"), c.get("degree_level"), c.get("country"), c.get("official_url")]
            for c in candidate_opportunities
            if c.get("university") or c.get("program") or c.get("degree_level")
        ],
    )

    professors_sheet = workbook.create_sheet("Professors")
    _write_sheet(
        professors_sheet,
        ["Name", "University", "Research Areas", "Profile URL"],
        [
            [c.get("professor_name"), c.get("university"), ", ".join(c.get("research_areas") or []), c.get("official_url")]
            for c in candidate_opportunities
            if c.get("professor_name")
        ],
    )

    funding_sheet = workbook.create_sheet("Funding")
    _write_sheet(
        funding_sheet,
        ["Title", "Funding Type", "Amount", "Deadline", "Official URL"],
        [
            [c.get("title"), c.get("funding_type"), c.get("funding_amount"), c.get("deadline"), c.get("official_url")]
            for c in candidate_opportunities
            if c.get("funding_type")
        ],
    )

    eligibility_sheet = workbook.create_sheet("Eligibility")
    _write_sheet(
        eligibility_sheet,
        ["Opportunity", "Eligibility", "Confidence", "Missing Requirements", "Warnings", "Explanation"],
        [
            [
                candidates_by_id.get(v["opportunity_id"], {}).get("title", v["opportunity_id"]),
                v.get("eligible"),
                v.get("confidence"),
                ", ".join(v.get("missing_requirements") or []),
                ", ".join(v.get("warnings") or []),
                v.get("explanation"),
            ]
            for v in eligibility_verdicts
        ],
    )

    ranked_sheet = workbook.create_sheet("Ranked Opportunities")
    research_by_id = {v["opportunity_id"]: v for v in research_match_verdicts}
    _write_sheet(
        ranked_sheet,
        ["Rank", "Opportunity", "Overall Score", "Research Match", "University", "Funding", "Official URL"],
        [
            [
                r.get("rank"),
                candidates_by_id.get(r["opportunity_id"], {}).get("title", r["opportunity_id"]),
                r.get("overall_score"),
                research_by_id.get(r["opportunity_id"], {}).get("overall_match"),
                candidates_by_id.get(r["opportunity_id"], {}).get("university"),
                candidates_by_id.get(r["opportunity_id"], {}).get("funding_type"),
                candidates_by_id.get(r["opportunity_id"], {}).get("official_url"),
            ]
            for r in ranked_opportunities
        ],
    )

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
