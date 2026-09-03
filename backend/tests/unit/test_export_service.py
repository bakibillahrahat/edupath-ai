from __future__ import annotations

import io

from openpyxl import load_workbook

from app.modules.ai_orchestration.service import build_workflow_workbook


def test_build_workflow_workbook_has_five_sheets_with_real_data():
    candidates = [
        {
            "id": "stanford-cs-phd", "title": "PhD in Computer Science", "university": "Stanford University",
            "program": "PhD", "degree_level": "PhD", "country": "USA", "professor_name": None,
            "research_areas": ["AI"], "funding_type": "Fully Funded", "funding_amount": None,
            "deadline": None, "official_url": "https://cs.stanford.edu",
        },
        {
            "id": "dr-smith", "title": "Dr. Smith", "university": "MIT", "program": None, "degree_level": None,
            "country": "USA", "professor_name": "Dr. Smith", "research_areas": ["ML"], "funding_type": None,
            "funding_amount": None, "deadline": None, "official_url": "https://mit.edu/smith",
        },
    ]
    eligibility = [
        {"opportunity_id": "stanford-cs-phd", "eligible": "likely_eligible", "confidence": 0.9, "missing_requirements": [], "warnings": [], "explanation": "Meets GPA."},
    ]
    research_match = [
        {"opportunity_id": "stanford-cs-phd", "overall_match": 0.85, "interest_overlap": 0.9, "technical_overlap": 0.8, "experience_alignment": 0.8, "program_alignment": 0.9, "explanation": "Strong fit."},
    ]
    ranked = [
        {"opportunity_id": "stanford-cs-phd", "overall_score": 0.82, "score_breakdown": {"research_match": 0.25}, "rank": 1},
    ]

    workbook_bytes = build_workflow_workbook(
        candidate_opportunities=candidates, eligibility_verdicts=eligibility,
        research_match_verdicts=research_match, ranked_opportunities=ranked,
    )

    workbook = load_workbook(io.BytesIO(workbook_bytes))
    assert workbook.sheetnames == ["Universities", "Professors", "Funding", "Eligibility", "Ranked Opportunities"]

    universities = list(workbook["Universities"].iter_rows(values_only=True))
    assert universities[0] == ("Title", "University", "Program", "Degree Level", "Country", "Official URL")
    assert universities[1][0] == "PhD in Computer Science"

    professors = list(workbook["Professors"].iter_rows(values_only=True))
    assert professors[1][0] == "Dr. Smith"

    funding = list(workbook["Funding"].iter_rows(values_only=True))
    assert funding[1][0] == "PhD in Computer Science"
    assert funding[1][1] == "Fully Funded"

    eligibility_rows = list(workbook["Eligibility"].iter_rows(values_only=True))
    assert eligibility_rows[1][0] == "PhD in Computer Science"  # resolved via candidates_by_id, not the raw id
    assert eligibility_rows[1][1] == "likely_eligible"

    ranked_rows = list(workbook["Ranked Opportunities"].iter_rows(values_only=True))
    assert ranked_rows[1][0] == 1
    assert ranked_rows[1][1] == "PhD in Computer Science"
    assert ranked_rows[1][3] == 0.85  # research match pulled in from research_match_verdicts


def test_build_workflow_workbook_handles_empty_data_without_crashing():
    workbook_bytes = build_workflow_workbook(
        candidate_opportunities=[], eligibility_verdicts=[], research_match_verdicts=[], ranked_opportunities=[],
    )
    workbook = load_workbook(io.BytesIO(workbook_bytes))
    assert workbook.sheetnames == ["Universities", "Professors", "Funding", "Eligibility", "Ranked Opportunities"]
    for sheet_name in workbook.sheetnames:
        rows = list(workbook[sheet_name].iter_rows(values_only=True))
        assert len(rows) == 1  # header only
